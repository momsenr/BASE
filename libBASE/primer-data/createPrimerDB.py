#!/usr/bin/python
import argparse
import openpyxl
import sys

from libBASE.primer import primer_IGHV, primer_IGHJ, primer_IGKV, primer_IGKJ, primer_IGLV

#####Parse command line arguments
parser = argparse.ArgumentParser(description='Load an prefilled .xls file, write an python-includable primer list. For backwards compatibility, the new primer list is compared agains libBASE/primer.py')
parser.add_argument('input', action='store', help='input file')
parser.add_argument('output', action='store', help='output file')

parser.add_argument('--gene', action='store', help='column where the gene name is found')
parser.add_argument('--primer', action='store', help='column where the kchainrimer is found')

args = parser.parse_args()

if(args.input==args.output):
    sys.exit("input file is output file. Please do not do that. Aborting ...")

try:
    workbook = openpyxl.load_workbook(args.input)
except FileNotFoundError:
        sys.exit("OOPS! File " + args.input + " not found! Aborting...")
except ValueError as my_err:
        sys.exit("OOPS! An error occured while parsing " + args.input +". " + str(my_err) +". Aborting ...")
except OSError as my_err:
        sys.exit("OOPS! An error occured while parsing " + args.input +". " + str(my_err) +". Maybe the wrong filetype? Aborting ...")

ws=workbook.active

primerdict_new={}
primerdict_old={}
primerdict_new['HV']={'N/A':''}
primerdict_new['HJ']={'N/A':''}
primerdict_new['KV']={'N/A':''}
primerdict_new['KJ']={'N/A':''} 
primerdict_new['LV']={'N/A':''} 
primerdict_new['LJ']={'N/A':''}

primerdict_old['HV']=primer_IGHV
primerdict_old['HJ']=primer_IGHJ
primerdict_old['KV']=primer_IGKV
primerdict_old['KJ']=primer_IGKJ
primerdict_old['LV']=primer_IGLV
primerdict_old['LJ']={}


primer_switched={}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    name=row[1].value
    gene_class=row[0].value
    primer=row[3].value

    if(name==None):
        continue
    primerdict_new[gene_class][name]=primer
    if(row[4].value!=None):#the primer was changed
        primer_switched[gene_class + " " + name]=row[4].value

for gene_class in ['HV','HJ','LV','KJ','LV','LJ']:
    for k,v in list(primerdict_old[gene_class].items()):
        try:
            if (primerdict_new[gene_class][k]!=v):
                try:
                    if(v==primer_switched[gene_class + " " + k]):#The primer change was correctly annotated in the excel file
                        continue
                    else:
                        print("The primer for " + gene_class + " " +k + " changed from " + v + " to " + primerdict_new[gene_class][k] + ". According to the raw excel file, the old primer was " + primer_switched[gene_class + " " + k])
                except: 
                    print("The primer for " + gene_class + " " +k + " changed from " + v + " to " + primerdict_new[gene_class][k] + ". This was not noted in the raw excel file.")
        except KeyError:
            print("Warning: There used to be a primer for the gene " + gene_class + " " +k + " (" + v + "), but there is no primer in the new list!")


preamble="""
#!/bin/python
\"\"\"
(c) 2019 Momsen Reincke <mail@momsenreincke.de> and Jakob Kreye <jakob.kreye@charite.de>

This file has been automatically generated using createPrimerDB.py
    
It defines the five dictionaries primer_IGHV, primer_IGHJ, primer_IGLV, primer_IGKV, primer_IGKJ. They contain the primer corresponding to certain IgG genes. 
    
    Example:
    The 5\' primer corresponding to IGH1-2*02 can be found in the dictionary primer_IGHV under the key '1-2*02'
    (the primer has the name H5 2/3-1). Analogously, the 3\' primer corresponding to IGK1-9*01 can be found 
    in the dictionary primer_IGKJ under the key 1-9*01 (the primer has the name K5 3-2)
\"\"\"
"""


f = open(args.output,"w")
f.write(preamble + "\n")
f.write("primer_IGHV=" + str(primerdict_new['HV'] ))
f.write('\n')
f.write("primer_IGHJ=" + str(primerdict_new['HJ'] ))
f.write('\n')
f.write("primer_IGKV=" + str(primerdict_new['KV'] ))
f.write('\n')
f.write("primer_IGKJ=" + str(primerdict_new['KJ'] ))
f.write('\n')
f.write("primer_IGLV=" + str(primerdict_new['LV'] ))
f.write('\n')
f.write("primer_IGLJ=" + str(primerdict_new['LJ'] ))
f.close() 
