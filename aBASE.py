#!/usr/bin/python
import argparse
import openpyxl
from libBASE.libBASE import SequenceFile
from libBASE.libBASE import updateExcelRow
from libBASE.libBASE import exportDict
import sys
from openpyxl.utils.cell import get_column_letter

import contextlib, os

###helper
def createExportDict(ws, begin, end):
    """returns a dict with keys='what should go in this column' and values='excel column name', created from worksheet ws
    """
    output_dict=dict()
    for row in ws[begin:end]:
        for cell in row:
            #190328: this used to be 
            #output_dict[cell.value]=cell.column
            #apparently now we have to use get_column_letter to convert the number into a letter
            output_dict[cell.value]=get_column_letter(cell.column)
    return output_dict 

def createCloningDict(ws, begin, end):
    """returns a dict with keys='what should go in this column' and values='excel column name', created from worksheet ws
    """
    cloning_output_dict=dict()
    for row in ws[begin:end]:
        for cell in row:
            cloning_output_dict[cell.value]=get_column_letter(cell.column)
    return cloning_output_dict 

#####Parse command line arguments
parser = argparse.ArgumentParser(description='aBase automatically igblasts sequencing reads and gives cloning suggestions and immunological annotations.')
parser.add_argument('input', action='store', help='input file')
parser.add_argument('output', action='store', help='output file')
parser.add_argument('--dataprefix', action='store',help='prefix for the sequence file filenames. this can be a directory name')
parser.add_argument('--hchain', action='store', help='column where the heavy chain is found')
parser.add_argument('--heavykeys', action='store', help='columns where the heavy chain output should be written to.')
parser.add_argument('--kappakeys', action='store', help='columns where the kappa chain output should be written to.')
parser.add_argument('--lambdakeys', action='store', help='columns where the lambda chain output should be written to.')
parser.add_argument('--identifier', action='store', help='Two columns where the identifier information for each mAb can be found (patient ID and mAb ID.)')
parser.add_argument('--cloningkeys', action='store', help='Five columns where the cloning recommendation output should be written to.')
parser.add_argument('--kchain', action='store', help='column where the kappa chain is found')
parser.add_argument('--lchain', action='store', help='column where the lambda chain is found')
parser.add_argument('--overwrite', action='store_true', help='overwrite')

args = parser.parse_args()

if(args.input==args.output):
    sys.exit("input file is output file. Please do not do that. Aborting ...")

try:
    #2019-03-28: keep_vba=True was added since the file could not be saved else (see: https://bitbucket.org/openpyxl/openpyxl/issues/766/workbook-cannot-saved-twice)
    workbook = openpyxl.load_workbook(args.input, keep_vba=True)
except FileNotFoundError:
        sys.exit("File " + args.input + " not found! Aborting...")
except ValueError as my_err:
        sys.exit("OOPS! An error occured while parsing " + args.input +". " + str(my_err) +". Aborting ...")
except OSError as my_err:
        sys.exit("OOPS! An error occured while parsing " + args.input +". " + str(my_err) +". Maybe the wrong filetype? Aborting ...")

ws=workbook.active


chains={}
begin={}
#analyzed_chains={}
end={}
columndict={}

if(args.hchain is not None):

    if(args.hchain.find(":")==-1):#we suppose this a single cell then
        args.hchain=args.hchain+":"+args.hchain

    chains['H']=workbook.active[args.hchain]
    if(args.heavykeys is not None):
        begin=args.heavykeys.split(":")[0]
        end=args.heavykeys.split(":")[1]
        columndict['H']=createExportDict(ws,begin,end)
        if('Comment' not in columndict['H'].keys()):
            sys.exit("No column titled 'Comment' in the fields specified by --heavykeys. Exiting...")

    else:
        print("--heavykeys not set. No heavy chain data will be written")

if(args.kchain is not None):
    if(args.kchain.find(":")==-1):#we suppose this a single cell then. The following line converts Z4 to Z4:Z4
        args.kchain=args.kchain+":"+args.kchain
    chains['K']=workbook.active[args.kchain]
    if(args.kappakeys is not None):
        begin=args.kappakeys.split(":")[0]
        end=args.kappakeys.split(":")[1]
        columndict['K']=createExportDict(ws,begin,end)
        if('Comment' not in columndict['K'].keys()):
            sys.exit("No column titled 'Comment' in the fields specified by --kappakeys. Exiting...")
    else:
        print("--kappakeys not set. No lambda chain data will be written")

if(args.lchain is not None):
    if(args.lchain.find(":")==-1):#we suppose this a single cell then. The following line converts Z4 to Z4:Z4
        args.lchain=args.lchain+":"+args.lchain
    chains['L']=workbook.active[args.lchain]
    if(args.lambdakeys is not None):
        begin=args.lambdakeys.split(":")[0]
        end=args.lambdakeys.split(":")[1]
        columndict['L']=createExportDict(ws,begin,end)
        if('Comment' not in columndict['L'].keys()):
            sys.exit("No column titled 'Comment' in the fields specified by --lambdakeys. Exiting...")
    else:
        print("--lambdakeys not set. No lambda chain data will be written")

if(args.cloningkeys is not None):
    begin=args.cloningkeys.split(":")[0]
    end=args.cloningkeys.split(":")[1]
    cloning_keys_dict=createCloningDict(ws,begin,end)
else:
    print("--cloningkeys not set. No cloning reccomendation will be written")

if(args.identifier is not None):
    patient_identifier_column=args.identifier.split(",")[0]
    mAb_identifier_column=args.identifier.split(",")[1]
else:
    print("--identifier not set. No clone-ID name will be written")

cloning_mAbs={}

for ct in chains.keys(): # ct is a chaintype, i.e. H, K or L
    ###chains['H']] is loaded by chains['H']=workbook.active[args.heavy]. if args.heavy=Z (a whole row),
    ###a list is returned, but if args.heavy=Z4:Z240 (for example..), then a tuple is returen (?!?).
    ### thats why we have to iterate over seq, instead of seq###
    for active_cell, in chains[ct]:
        if active_cell.value is None:
            continue
        
        #check if this line has already been analyzed - and skip, args.overwrite is not set
        if(active_cell.value is not None and args.overwrite is False):
            if(ws[columndict[ct]["Confirmation"]+str(active_cell.row)].value!=None): 
                print(ws[columndict[ct]["Confirmation"]+str(active_cell.row)].value + " has already been analyzed.")
                continue

        filename=active_cell.value
        
        #21.03.21 the following line is a workaround for the inconsistent naming scheme of Eurofins
        filename=filename.replace("-","_")
        
        if(args.dataprefix is not None):
            filename=args.dataprefix+str(filename)+".ab1"
        try:
            my_ps=SequenceFile(filename)

            if(my_ps.successfullyParsed==False):
                ws[columndict[ct]['Comment']+str(active_cell.row)]=my_ps.comment 
                ws[columndict[ct]['QV']+str(active_cell.row)]=my_ps.mean_phred_quality
                ws[columndict[ct]['Confirmation']+str(active_cell.row)]="to be confirmed"
                ws[columndict[ct]['Function']+str(active_cell.row)]="BQ"
                ws[columndict[ct]['RL']+str(active_cell.row)]=my_ps.len 
            elif(my_ps.chain_type is not ct):
                ws[columndict[ct]['Comment']+str(active_cell.row)]=my_ps.comment+" "+filename + " has chain type " + my_ps.chain_type
                ws[columndict[ct]['QV']+str(active_cell.row)]=my_ps.mean_phred_quality
                ws[columndict[ct]['RL']+str(active_cell.row)]=my_ps.len 
                ws[columndict[ct]['Confirmation']+str(active_cell.row)]="to be confirmed"
                ws[columndict[ct]['Function']+str(active_cell.row)]="BQ"
            else:
                #TODO: deprecate updateExcelRow by moving the code of updateExcelRow here
                updateExcelRow(workbook,active_cell.row,columndict[ct],my_ps)
                
                #record cloning informationin cloning_mAbs
                if(args.cloningkeys is not None):
                    ed=exportDict(my_ps)
                    if(ed["5' Primer"].find(ct)!=-1 and ed["3' Primer"].find(ct)!=-1):
                        if(ed["Function"]=="Y"):
                            try:
                                cloning_mAbs[active_cell.row]+=ct
                            except:
                                cloning_mAbs[active_cell.row]=ct
                        else:
                            try:
                                cloning_mAbs[active_cell.row]+=ct+"*"
                            except:
                                cloning_mAbs[active_cell.row]=ct+"*"
        except FileNotFoundError:
            try:
                ws[columndict[ct]['Function']+str(active_cell.row)]="BQ - file not found"
                ws[columndict[ct]['Comment']+str(active_cell.row)]="File "+str(filename)+" not found."
            except:
                sys.exit("OOPS! File " + filename + " not found! Also, either 'Comment' or 'Function' column was not found. Please make sure there are columns with that name in the range specified.")
        except ValueError as my_err:
            sys.exit("OOPS! An error occured while parsing " + filename +". " + str(my_err) +". Aborting ...")
        except OSError as my_err:
            sys.exit("OOPS! An error occured while parsing " + filename +". " + str(my_err) +". Maybe the wrong filetype? Aborting ...")

if(args.heavykeys is None or args.lambdakeys is None or args.kappakeys is None):
    sys.exit("Please set keys")

if(args.cloningkeys is not None):
    for row in cloning_mAbs.keys():
        # we will only clone mAbs with heavy chains
        if(cloning_mAbs[row].find("H")==-1 or cloning_mAbs[row]=="H*" or cloning_mAbs[row]=="H" or cloning_mAbs[row]=="H*K*" or cloning_mAbs[row]=="H*L*" or cloning_mAbs[row]=="H*K*L*"):
            continue
        else:
            #in case of non-functional chains, we sometimes want to restrict the chains we're cloning
            #e.g. in case  of e.g. a functional H and L chain with a non-function K chain ("HK*L") we want to clone only H and L
            if(cloning_mAbs[row]=="HK*L"):
                cloning_mAbs[row]="HL"
            elif(cloning_mAbs[row]=="HKL*"):
                cloning_mAbs[row]="HK"
            elif(cloning_mAbs[row]=="H*K*L"):
                cloning_mAbs[row]="H*L"
            elif(cloning_mAbs[row]=="H*KL*"):
                cloning_mAbs[row]=="H*K"

            ws[cloning_keys_dict["cloning?"]+str(row)]=cloning_mAbs[row].replace("*","")

            #write non-functional chains to be cloned.
            #there is only one combination where two non-functional chains will be cloned, "HK*L*", this is handled in the first clause
            #else there is just one non-functional chain max
            if(cloning_mAbs[row]=="HK*L*"):
                ws[cloning_keys_dict["non functional chains"]+str(row)]= "K*L*" 
            elif(cloning_mAbs[row].find("*")!=-1):
                ws[cloning_keys_dict["non functional chains"]+str(row)]= cloning_mAbs[row][cloning_mAbs[row].find("*")-1] + "*" 
            
            if(args.identifier is not None):#write clone-IDs, e.g. 003-102
                ws[cloning_keys_dict["clone ID"]+str(row)]=str(ws[patient_identifier_column+str(row)].value) +"-"+ str(ws[mAb_identifier_column+str(row)].value)


# suppress "FutureWarning: The behavior of this method will change in future versions. Use specific 'len(elem)' or 'elem is not None' test instead." in 
with open(os.devnull, 'w') as devnull:
    with contextlib.redirect_stderr(devnull):
        workbook.save(args.output)

