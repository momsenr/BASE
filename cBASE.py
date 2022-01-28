#!/usr/bin/python
#v2022_01_28
import argparse
import openpyxl
from libBASE.libBASE import SequenceFile
from libBASE.libBASE import updateExcelRow
from libBASE.libBASE import exportDict
from Bio.Seq import Seq
import sys

from libBASE.libBASE import AlignPCRObject

from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, colors

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
yellowFill = PatternFill(start_color='ffff00',
                   end_color='ffff00',
                   fill_type='solid')
orangeFill = PatternFill(start_color='d1731e',
                   end_color='d1731e',
                   fill_type='solid')
lightgreenFill = PatternFill(start_color='92D050',
                   end_color='92D050',
                   fill_type='solid')
deepgreenFill = PatternFill(start_color='00B050',
                   end_color='00B050',
                   fill_type='solid')
purpleFill = PatternFill(start_color='8f22ba',
                   end_color='8f22ba',
                   fill_type='solid')
greyFill = PatternFill(start_color='b2b2b2',
                   end_color='b2b2b2',
                   fill_type='solid')
pinkFill = PatternFill(start_color='ff99ff',
                   end_color='ff99ff',
                   fill_type='solid')


#####Parse command line arguments
parser = argparse.ArgumentParser(description='cBase compares the sequencing data of plasmids and PCR reads on a nucleotide per nucleotide basis.')
parser.add_argument('input', action='store', help='input file')
parser.add_argument('output', action='store', help='output file')
parser.add_argument('--dataprefix', action='store',help='prefix for the sequence file filenames. this can be a directory name')
parser.add_argument('--pcr2read', action='store', help='column where the name of the sequencing file of the 2nd pcr is found')
parser.add_argument('--plasmidread', action='store', help='column where the name of the sequencing file of the plasmid is found')
parser.add_argument('--shmanalysis', action='store', help='columns where the somatic hypermutation analysis should be written to. If four columns (separated by a comma) instead of one are given, the software will also give a more detailed analysis of the somatic hypermutations of the pcr2 read, of the plasmid, and the idealized antibody.')
parser.add_argument('--manualanalysis', action='store', help='columns where the expression recommendation/manual analysis should be written')

args = parser.parse_args()

if(args.input==args.output):
    sys.exit("input file is output file. Please do not do that. Aborting ...")

try:
    workbook = openpyxl.load_workbook(args.input)
except FileNotFoundError:
        sys.exit("File " + args.input + " not found! Aborting...")
except ValueError as my_err:
        sys.exit("OOPS! An error occured while parsing " + args.input +". " + str(my_err) +". Aborting ...")
except OSError as my_err:
        sys.exit("OOPS! An error occured while parsing " + args.input +". " + str(my_err) +". Maybe the wrong filetype? Aborting ...")

ws=workbook.active

if(args.pcr2read is not None):
    if(args.pcr2read.find(":")==-1):#we suppose this is a single cell then
        args.pcr2read=args.pcr2read+":"+args.pcr2read
    pcr2read=workbook.active[args.pcr2read]
if(args.shmanalysis is not None):
    differential_analysis_column=args.shmanalysis[0]
    if(len(args.shmanalysis)>1):
            pcr2_shm_column=args.shmanalysis.split(",")[1]
            plasmid_shm_column=args.shmanalysis.split(",")[2]
            ideal_shm_column=args.shmanalysis.split(",")[3]
            
to_compare=[]

###chains['H']] is loaded by chains['H']=workbook.active[args.heavy]. if args.heavy=Z (a whole row),
###a list is returned, but if args.heavy=Z4:Z240 (for example..), then a tuple is returen (?!?).
### thats why we have to iterate over seq, instead of seq###
output=""
green=False
yellow=False
for seq, in pcr2read:
    ###TODO: error handling!
    if seq.value is None:
        continue
    filename_pcr2=seq.value
    filename_pcr2=args.dataprefix+str(filename_pcr2)+".ab1"
    
    #the following line is a workaround for the inconsistent naming scheme of Eurofins
    filename_pcr2=filename_pcr2.replace("-","_")
    
    try:
        pcr2=SequenceFile(filename_pcr2) 
        
        filename_plasmid=ws[args.plasmidread+str(seq.row)].value
        filename_plasmid=args.dataprefix+str(filename_plasmid)+".ab1"
    
        #the following line is a workaround for the inconsistent naming scheme of Eurofins
        filename_plasmid=filename_plasmid.replace("-","_")
        
        print("Comparing: " + filename_pcr2 + " " + filename_plasmid)

        try:
            plasmid=SequenceFile(filename_plasmid)
            if(plasmid.successfullyParsed==True):
                aligned_Sequences=AlignPCRObject(pcr2,plasmid)
                output=aligned_Sequences.output
            else:
                output="BQ: Could not blast " + plasmid.filename + ". This could either be due to bad sequencing quality, or maybe because it's an empty vector? igblast complained: " + plasmid.comment
                
        except FileNotFoundError as err:
            output="FileNotFound"
            print(str(err))
        except:
            output="Uncaught exception. Please inform the author of this software and provide the ab1-file(s)." 
       
        output_cell=differential_analysis_column+str(seq.row)

        ws[output_cell]=output 
        
        #Here we color-code the cBASE output
        if(output=="0"):
            ws[output_cell].fill = deepgreenFill
            green=True
        elif(output.find("FileNotFound")!=-1):
            pass
        elif(output.find("Uncaught")!=-1 or output.find("BQ")!=-1 or output.find("empty")!=-1 or output.find("chain types differ!")!=-1):#"uncaught exception", "bad quality", "empty vector", "Diff HC/KC/LC" - these will be checked before we check for productivity of the chains
            ws[output_cell].fill = redFill
        elif(output.find("completely different chains")!=-1 or output.find("WARNING")!=-1 or output.find("non-functional")!=-1 or output.find("CAVE")!=-1 or output.find("do not match")!=-1):#chain types differ or likely mutation in primer region - manual analysis necessary or V/J genes do not match #remark 22.02.2021: this this elif clause needs to be before the rest, since the aligned_Sequences.D1 object might not exist if the chain types differ
            ws[output_cell].fill = orangeFill
        elif(aligned_Sequences.D1['productive'].lower()!='yes' and aligned_Sequences.D2['productive'].lower()!='yes'):
            ws[output_cell].fill = greyFill
        elif(aligned_Sequences.D1['productive'].lower()!='yes' and aligned_Sequences.D2['productive'].lower()=='yes'):
            ws[output_cell].fill = purpleFill
        elif(aligned_Sequences.D1['productive'].lower()=='yes' and aligned_Sequences.D2['productive'].lower()!='yes'):
            ws[output_cell].fill = redFill
        elif(output.find("index")!=-1 ):#index error #TODO orange or red?
            ws[output_cell].fill = redFill
        elif(output.find("nsSHM+")!=-1 or output.find("nsSHMchg")!=-1):
            ws[output_cell].fill = yellowFill
            yellow=True
            if(aligned_Sequences.total_nonsilent_mutations==3):
                ws[output_cell].fill = orangeFill
                yellow=False
            elif(aligned_Sequences.total_nonsilent_mutations>3): 
                ws[output_cell].fill = redFill
                yellow=False
        else:
            ws[output_cell].fill = lightgreenFill
            green=True


        if(len(args.shmanalysis)>1):
            try:
                if(aligned_Sequences.pcr1.successfullyParsed==True):
                    ed1=exportDict(aligned_Sequences.pcr1)
                    ws[pcr2_shm_column+str(seq.row)]=str(ed1['SHM'])
                else:
                    ws[pcr2_shm_column+str(seq.row)]="n/a"
            except:
                ws[pcr2_shm_column+str(seq.row)]="n/a"
            try:
                if(aligned_Sequences.pcr2.successfullyParsed==True):
                    ed2=exportDict(aligned_Sequences.pcr2)
                    ws[plasmid_shm_column+str(seq.row)]=str(ed2['SHM'])
                else:
                    ws[plasmid_shm_column+str(seq.row)]="n/a"
            except:
                ws[plasmid_shm_column+str(seq.row)]="n/a"
            try:
                if(aligned_Sequences.pcr2.successfullyParsed==True and aligned_Sequences.pcr1.successfullyParsed==True):
                    ws[ideal_shm_column+str(seq.row)]=aligned_Sequences.number_of_shm_v_gene_ideal
                else:
                    ws[ideal_shm_column+str(seq.row)]="n/a"
            except:
                ws[ideal_shm_column+str(seq.row)]="n/a"
                
        if(green is not True):
            print(output)
        
        if(args.manualanalysis is not None):
            ws[args.manualanalysis+str(seq.row)].alignment=Alignment(horizontal='center')
            if(green == True):
                ws[args.manualanalysis+str(seq.row)].fill=deepgreenFill
                ws[args.manualanalysis+str(seq.row)]="OK"
            elif(yellow == True):
                ws[args.manualanalysis+str(seq.row)].fill=yellowFill
                ws[args.manualanalysis+str(seq.row)]=str(aligned_Sequences.total_nonsilent_mutations) + " SHM"
            else:
                ws[args.manualanalysis+str(seq.row)].fill=pinkFill
                ws[args.manualanalysis+str(seq.row)]="open"
            
        output=""
        aligned_Sequences=None
        green=False
        yellow=False
    except FileNotFoundError as err:
        print(str(err))
    except ValueError as my_err:
            sys.exit("OOPS! An error occured while parsing " + filename_pcr2 +". " + str(my_err) +". Aborting ...")
    except OSError as my_err:
            sys.exit("OOPS! An error occured while parsing " + filename_pcr2 +". " + str(my_err) +". Maybe the wrong filetype? Aborting ...")

workbook.save(args.output)
